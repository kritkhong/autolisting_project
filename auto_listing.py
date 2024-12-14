
import PIL
import re
import base64
import json
import shutil
import math
import imagehash
from collections import deque
from natsort import natsorted
from datetime import datetime
from pathlib import Path
from paddleocr import PaddleOCR, draw_ocr
from PIL import Image, ImageDraw
from dotenv import dotenv_values
from openai import OpenAI
from openpyxl import load_workbook


# Adjustable variable
img_size = 500
# stock_code = f'{stock_name}{stock_no:04}'
stock_name = datetime.now().strftime(r'%d%m%y')
stock_no = 1
item_count = 10


# User's input
while (True):
    dir_path = Path(input('Specify full directory address of target folder: '))
    if dir_path.exists():
        break
    else:
        print('Path doesn\'t exist. Please check the address or path format or any typos.')
code_prefix = input(
    'Please specify the LETTER(s) code for more accuracy [if not just press \'Enter\']: ')


# resize entire folder of image and save to jpg format (acceptable format for chatGpt) return path of the result folder
def batch_resize_imgs(folder_path: Path, img_size: int) -> Path:
    result_folder = folder_path / \
        ('result_'+datetime.now().strftime(r'%d%m%y_%H%M%S'))
    result_folder.mkdir(exist_ok=True)
    for item in folder_path.iterdir():
        try:
            im = Image.open(item)
        except (PIL.UnidentifiedImageError, IsADirectoryError):
            continue
        ratio = min(im.size)/img_size if min(im.size) > img_size else 1
        size = (int(im.width / ratio), int(im.height / ratio))
        resized_img = im.resize(size)
        output_name = f'{result_folder}/{item.stem}.jpg'
        resized_img.save(output_name)
    return result_folder


def create_caption(item: Path, code_list: list, dummy=False) -> list:
    output = {}
    if len(code_list) == 1:
        caption = single_caption(item, dummy)
        output[code_list[0]] = caption

    elif len(code_list) > 1:
        captions = multi_captions(item, code_list, dummy)
        for code in code_list:
            try:
                output[code] = captions[code]
            except:
                output[code] = '*** Caption Create Error ***'
    return output


def read_code_price(item: Path, code_prefix: str, ocr, draw=False) -> list:
    text_read = ocr.ocr(f'{item}', cls=False)
    text_read = text_read[0]
    code_list = []
    price_list = []
    if (text_read):
        # if code prefix is not specified make it to A-Z
        code_prefix = 'A-Z' if code_prefix == '' else code_prefix.upper()
        # Regex explain: 1 Alphabet follow by 01-09 , 10-9999
        # sometimes ocr got confused between O(letter) and 0(number)
        code_regex = r'([' + f'{code_prefix}' + \
            r']([0O][1-9]|[1-9][O\d]{1,3}))'
        # Regex explain: (start of str / = / space)(3-5 digits number not start with 0)
        price_regex = r'([1-9]\d{2,3})'
        for line in text_read:
            box_loc = line[0]
            str = line[1][0]
            # Extract code
            while (True):
                match = re.search(code_regex, str)
                if (match == None):
                    break
                # append the code if have letter O replace with 0
                code_list.append([box_loc, re.sub('O', '0', match.group(0))])
                # trim that code out of the string
                str = str[:match.start()] + ' ' + str[match.end():]
            # Extract price
            while (True):
                match = re.search(price_regex, str)
                if (match == None):
                    break
                # append the code if have letter O replace with 0
                price_list.append([box_loc, match.group()])
                # trim that code out of the string
                str = str[:match.start()] + ' ' + str[match.end():]
    if draw and price_list:
        for box, price in price_list:
            draw_box(box, 'PRICE', (3, 252, 15), item)

    # one code
    # multiple code
    output = []
    for box_c, code in code_list:
        if draw:
            draw_box(box_c, 'CODE', (3, 252, 252), item)
        if price_list:
            x1 = [p[0] for p in box_c]
            y1 = [p[1] for p in box_c]
            code_center = [sum(x1)/4, sum(y1)/4]
            min_dist = float('inf')
            pair_price = []
            for box_p, price in price_list:
                x2 = [p[0] for p in box_p]
                y2 = [p[1] for p in box_p]
                price_center = [sum(x2)/4, sum(y2)/4]
                distance = math.dist(code_center, price_center)
                if distance < min_dist:
                    min_dist = distance
                    pair_price = [price_center, price]
            if draw:
                draw_line(code_center, pair_price[0], item)
            output.append([code, pair_price[1]])
        else:
            output.append([code, None])

    # return in list of [code, price]
    # [ [code1,price1], [code2,price2], ... ]
    return output


def draw_box(box: list, text: str, color: tuple, img: Path):
    with Image.open(img) as im:
        draw = ImageDraw.Draw(im)
        draw.text([box[0][0], box[0][1]-15], text, fill=color,
                  font_size=12, stroke_width=1, stroke_fill=(0, 0, 0))
        draw.line(box[0]+box[1], fill=color, width=1)
        draw.line(box[1]+box[2], fill=color, width=1)
        draw.line(box[2]+box[3], fill=color, width=1)
        draw.line(box[3]+box[0], fill=color, width=1)
        im.save(img)
        return 1


def draw_line(a: list, b: list, img: Path):
    with Image.open(img) as im:
        draw = ImageDraw.Draw(im)
        draw.line(a+b, fill=(252, 3, 227, 128), width=1)
        im.save(img)
        return 1


def single_caption(item: Path, dummy: bool) -> str:
    ##### FOR TESTING #####
    if dummy:
        return "--test caption--"
    #######################

    img_path = str(item)
    with open(img_path, "rb") as image_file:
        base64_image = base64.b64encode(image_file.read()).decode('utf-8')

    env_vars = dotenv_values('.env')

    client = OpenAI(api_key=env_vars['OPEN_AI_KEY'])
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": [
                    {
                        "type": "text",
                        "text": "Your duty is product captioning. \nMy products are mainly Disneyland and Disney's store products and some other store's product that fall in the same category such as gift ,souvenir that have cartoon or fantasy character design, sometime there are other kind of product too, you can caption this kind as you see proper.\nHow this work is\nI will send you the image of the product. You only return the caption phrase (in Thai language), not a sentence.\nThe caption phrase have a brief structure as below:\n1. Type of product\n2. Color of product (if have multiple indicates a few main ones)\n3. Design or print on product : this part can be what character, color of that character, pattern, color of pattern, special texture etc as you see proper.\n\nfor example: กระติกน้ำเก็บอุณหภูมิสีแดง รูป Mickey Mouse ใส่ชุดสีน้ำเงิน\nnotes:\n- Sometimes, there is a hint of product type in the picture as a text in Thai.\n- You can name character in English\n- We have some usual products, use these words: \n'พวง ตต ' is plush keychain (small doll with keyring), \n'ตต ' is plush doll (relatively bigger dool),\n'กระเป๋าเหรียญ' is small pocket with keyring, can be plush keyring with zip pocket\n"
                    }
                ]
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            },
        ],
        temperature=1,
        max_tokens=1000,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0,
        response_format={
            "type": "text"
        }
    )
    caption = response.choices[0].message.content
    print(caption)
    return caption


def multi_captions(item: Path, code_list: list, dummy: bool) -> dict:
    ##### FOR TESTING #####
    if dummy:
        test_dict = {}
        for code in code_list:
            test_dict[code] = "--test caption--"
        return test_dict
    #######################
    img_path = str(item)
    with open(img_path, "rb") as image_file:
        base64_image = base64.b64encode(image_file.read()).decode('utf-8')

    env_vars = dotenv_values('.env')

    client = OpenAI(api_key=env_vars['OPEN_AI_KEY'])
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": [
                    {
                        "type": "text",
                        "text": "Your duty is product captioning. \nMy products are mainly Disneyland and Disney's store products and some other store's product that fall in the same category such as gift ,souvenir that have cartoon or fantasy character design, sometime there are other kind of product too, you can caption this kind as you see proper.\nHow this work is\nI will send you the image of the product and list(array) contain product label code that label on the image.\nYou only return in raw JSON string of input product label code as a key and the caption phrase (in Thai language) as a value.\nIf my input is wrong, please explain the error in english uppercase in the caption.\nfor example:  input = ['A05','K034','A28'] \noutput = { \"A05\": caption#1, \"K034\": caption#2, \"A28\": \"NOT FOUND**\"}\nThe caption phrase have a brief structure as below:\n1. Type of product\n2. Color of product (if have multiple indicates a few main ones)\n3. Design or print on product : this part can be what character, color of that character, pattern, color of pattern, special texture etc as you see proper.\n\nfor example: กระติกน้ำเก็บอุณหภูมิสีแดง รูป Mickey Mouse ใส่ชุดสีน้ำเงิน\nnotes:\n- Sometimes, there is a hint of product type in the picture as a text in Thai.\n- You can name character in English\n- We have some usual products, use these words: \n'พวง ตต ' is plush keychain (small doll with keyring), \n'ตต ' is plush doll (relatively bigger dool),\n'กระเป๋าเหรียญ' is small pocket with keyring, can be plush keyring with zip pocket"
                    }
                ]
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    },
                    {
                        "type": "text",
                        "text": str(code_list)
                    }
                ]
            },
        ],
        temperature=1,
        max_tokens=1000,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0,
        response_format={
            "type": "text"
        }
    )
    str_response = json_trim(response.choices[0].message.content)
    print(str_response)
    return json.loads(str_response)


def json_trim(string: str) -> str:
    regex = r'[{\[].*[}\]]'
    match = re.search(regex, string, re.DOTALL)
    return match.group()


def is_similar_img(img1: Path, img2: Path) -> bool:
    hash0 = imagehash.average_hash(Image.open(img1))
    hash1 = imagehash.average_hash(Image.open(img2))
    cutoff = 5  # maximum bits that could be different between the hashes.
    return ((hash0 - hash1) < cutoff)


def img_rename(img: Path, code_list: list) -> Path:
    list = natsorted(code_list)
    if (list[0]):
        new_img = img.parent / ('_'.join(list) + img.suffix)
    while (new_img.exists()):
        if (str(new_img) != str(img)):  # other file having this name(code) already exist
            if (is_similar_img(new_img, img)):  # same images
                new_img = new_img.parent / \
                    (new_img.stem + ' copy' + new_img.suffix)
            else:  # user use same code for different images for any reason
                new_img = new_img.parent / \
                    (new_img.stem + '*' +
                     new_img.suffix)  # put * for renaming file not overwrite other file with same code

        else:  # rename to its own name
            break
    img.rename(str(new_img))
    return new_img


## MAIN FUCTION ##

ocr = PaddleOCR(lang='en', show_log=False)
##########################################
## force PaddleOCR to use det_lang='ml' ##
## see paddleocr.py line 646            ##
##########################################

# batch_resize_imgs return path of the result folder
imgs_dir = batch_resize_imgs(dir_path, img_size)
listing = []
code_repeat = {}
for img in imgs_dir.iterdir():
    print(img)
    if (img.suffix == '.xlsx'):
        continue

    cp_list = read_code_price(img, code_prefix, ocr)
    code_list = [code for code, price in cp_list]

    # image have code(s) --> rename img, create caption(s), put info in listing
    if cp_list:
        img = img_rename(img, code_list)
        if ' copy' not in img.stem:  # non duplicate img file
            # <---- create dummy caption for testing change to False in real work
            caption_dict = create_caption(img, code_list, dummy=False)
            for code, price in cp_list:
                if (code not in code_repeat.keys()):
                    code_repeat[code] = 1
                    new_code = code
                else:
                    code_repeat[code] += 1
                    new_code = code + '_' + str(code_repeat[code])
                    new_file_path = img.parent / \
                        img.name.replace('*', '').replace(code, new_code)
                    img = img.rename(str(new_file_path))

                listing.append([new_code, caption_dict[code], price])
    # image doesn't contain code --> ignore
    else:
        continue

# EXCEL PART
path_temp_xls = Path('./stock_template.xlsx')
path_result_xls = imgs_dir / \
    ('stock_' + datetime.now().strftime(r'%d%m%y_%H%M%S') + '.xlsx')
shutil.copy(path_temp_xls, path_result_xls)

wb = load_workbook(path_result_xls)
ws = wb.active

listing = natsorted(listing, key=lambda y: y[0])

for info in listing:
    # stock_name is hard coded for sake of testing you can make it user input
    stock_code = f'{stock_name}{stock_no:04}'
    # A:stock_code B:sale_code C:caption E:count(default=10) F:price
    row_val = {
        'A': stock_code,
        'B': info[0],
        'C': info[1],
        'E': item_count,
        'F': info[2],
    }
    ws.append(row_val)
    stock_no += 1

    # check_info = {
    #     'row': ws.max_row,
    #     'col': [],
    #     'img': img_name
    # }
    # # check to see if this row need manual work
    # if row_val['B'] == None:
    #     check_info["col"].append('B')
    # if row_val['F'] == None or int(row_val['F']) > price_check:
    #     check_info["col"].append('F')
    # # if manual needed put it queue
    # if len(check_info["col"] > 0):
    #     check_needed.append(check_info)
    #     # TODO
    #     # if GUI not running ----> run it
    #     gui_check()
    #     # let user check while automation is running


wb.save(path_result_xls)
